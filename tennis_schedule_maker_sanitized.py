#Version 5 To modify your code so that all the releases are printed together, followed by all the reminders, you can separate the "release" section from the "reminder" section in two loops. This way, the output will first have all the releases, then all the reminders.

import openpyxl
from datetime import datetime

# Load the Excel spreadsheet
workbook = openpyxl.load_workbook('tennis_sched.xlsx')
sheet = workbook.active

# Prompt the user to enter the starting week number (1-based index)
print("\n" + "=" * 50)
print("🎾  Tennis Schedule Generator")
print("=" * 50)
starting_week = int(input("Enter the starting week number (1, 2, 3, ...): "))

# Determine the total number of weeks based on the row count and structure of your sheet
# Assuming each week starts every 5 rows
total_weeks = sheet.max_row // 5

# Output to a text file
output_filename = 'schedule_output.txt'
with open(output_filename, 'w') as output_file:
    # First loop for the releases
    for match_number in range(starting_week, total_weeks + 1):
        # Calculate the starting row based on the week number
        starting_row = 2 + (match_number - 1) * 5  # Each week starts every 5 rows

        # Calculate the cell references for D2 and E2 based on the week number
        d_cell_reference = f'D{2 + (match_number - 1) * 5}'
        e_cell_reference = f'E{2 + (match_number - 1) * 5}'

        # Get the match_info from cells D2 and E2 for the current week
        match_info = f"{sheet[d_cell_reference].value} || {sheet[e_cell_reference].value}"

        # Define the cell to skip based on the week number
        skip_cell_value = f'C{starting_row + 5}'

        # Read data from cells A3:B5 and C3:C5
        data = []
        for row in sheet.iter_rows(min_row=starting_row, max_row=starting_row + 3, min_col=1, max_col=3, values_only=True):
            if row[2] == skip_cell_value:
                continue  # Skip rows with the same value as skip_cell_value
            data.append(row)

        # # Write the release section for each match
        # Write the release section for each match
        # Write the release section for each match
        output_file.write("\n" + "=" * 32 + "\n")
        output_file.write(f"📢 SCHEDULE RELEASE\n")
        output_file.write(f"Match {match_number}: {match_info}\n")
        output_file.write("-" * 60 + "\n")

        for row in data:
            output_file.write(f"{row[0]}: {row[1]}")
            if row[2] is not None:
                output_file.write(f" - {row[2]}")
            output_file.write("\n")

        output_file.write("\n")
        output_file.write("-" * 60 + "\n")
        output_file.write("📌 Please keep status updated via text or at the sheet:\n")
        output_file.write(
            "YOUR_SPREADSHEET_URL_HERE\n")
        output_file.write("=" * 32 + "\n\n")

       
    # Second loop for the reminders
    for match_number in range(starting_week, total_weeks + 1):
        # Calculate the starting row based on the week number
        starting_row = 2 + (match_number - 1) * 5  # Each week starts every 5 rows

        # Calculate the cell references for D2 and E2 based on the week number
        d_cell_reference = f'D{2 + (match_number - 1) * 5}'
        e_cell_reference = f'E{2 + (match_number - 1) * 5}'

        # Get the match_info from cells D2 and E2 for the current week
        match_info = f"{sheet[d_cell_reference].value} || {sheet[e_cell_reference].value}"

        # Define the cell to skip based on the week number
        skip_cell_value = f'C{starting_row + 5}'

        # Read data from cells A3:B5 and C3:C5
        data = []
        for row in sheet.iter_rows(min_row=starting_row, max_row=starting_row + 3, min_col=1, max_col=3, values_only=True):
            if row[2] == skip_cell_value:
                continue  # Skip rows with the same value as skip_cell_value
            data.append(row)

      
        # Write the reminder section for each match
        output_file.write("\n" + "=" * 36 + "\n")
        output_file.write(f"⏰ SCHEDULE REMINDER\n")
        output_file.write(f"Match {match_number}: {match_info}\n")
        output_file.write("-" * 60 + "\n")

        for row in data:
            output_file.write(f"{row[0]}: {row[1]}")
            if row[2] is not None:
                output_file.write(f" - {row[2]}")
            output_file.write("\n")

        output_file.write("\n")
        output_file.write("📩 Please confirm.\n")
        output_file.write("=" * 36 + "\n\n")

